# -*- coding: utf-8 -*-
"""Importa base de apostas exportada de outro sistema (.xlsx, 15 colunas).

Genérico por --dono: mesmo formato de export, um dono por planilha. Usado para
Diogo (base própria) e Primo (operador do Diogo) — ambos exportaram do mesmo
sistema. Faz o de-para ao padrão do projeto (MASTER_*): esporte e casa ao canon,
mercado→categoria (27 canônicas), Status→Resultado (W/L/V/HW/HL), tipster com
fusão de duplicatas de caixa, e sanitiza caracteres de controle do texto. P/L é
derivado (não persistido).

Modelo de conta (espelha Feca/Lava e Fatuch/LavaFatuch): Diogo é DONO com base
própria sob dono='Diogo'; Primo é operador dele, base sob dono='Primo'. Diogo
enxerga a base do Primo via "ver como".

A assinatura é computada localmente, idêntica a repository._assinatura, e a
escrita é um bulk INSERT numa única transação, com retry — robusto contra a
queda do proxy público do Railway.

Uso:
    python scripts/import_apostas_xlsx.py --dono Diogo --xlsx "C:\\...\\diogo.xlsx" --dry
    python scripts/import_apostas_xlsx.py --dono Primo --xlsx "C:\\...\\primo.xlsx" --go
"""
import argparse
import asyncio
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
PARCEIRO = 'Padrão'          # sem coluna de titular nessas planilhas
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}


# ---------- sanitização de texto (remove control chars do export) ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')       # C0 + DEL (inclui \n \r \t \x0c)
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')      # escape literal do openpyxl (_x000C_)


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))                # tira o "_x000C_" que o export deixa literal
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa → canon (espelha _casa_display de app/main.py + data.js) ----------
_CASA_DISPLAY = {
    "BET365": "Bet365", "BETANO": "Betano", "BETFAIR": "Betfair",
    "BETNACIONAL": "Betnacional", "BOLSADEAPOSTA": "Bolsa de Aposta",
    "JOGODEOURO": "Jogo de Ouro", "KINGPANDA": "KingPanda", "KTO": "KTO",
    "LOTTU": "Lottu", "PINNACLE": "Pinnacle", "POLYMARKET": "Polymarket",
    "SUPERBET": "Superbet", "VITORIABET": "Vitória Bet",
    # casas trazidas pelas bases do Diogo/Primo (display do dashboard data.js)
    "NOVIBET": "Novibet", "BINGOPLUS": "BingoPlus", "VAIDEBET": "VaiDeBet",
    "BETMGM": "BetMGM", "ESPORTESDASORTE": "Esportes da Sorte",
    "SPORTINGBET": "SportingBet", "BETESPORTE": "BETesporte", "BETFAST": "Betfast",
}


def norm_casa(c: str) -> str:
    key = limpa(c).upper().replace(' ', '')
    return _CASA_DISPLAY.get(key, limpa(c).title())


# ---------- esporte → canon (MASTER_ESPORTES / mapa do dashboard) ----------
_ESPORTE_MAP = {
    'vários': 'Múltiplos', 'varios': 'Múltiplos',
    'futebol': 'Futebol', 'tênis': 'Tênis', 'tenis': 'Tênis',
    'dardos': 'Dardos', 'vôlei': 'Vôlei', 'volei': 'Vôlei',
    'basquete': 'Basquete', 'e-sports': 'E-Sports', 'esports': 'E-Sports',
    'fórmula 1': 'Fórmula 1', 'formula 1': 'Fórmula 1', 'automobilismo': 'Fórmula 1',
    'rugby': 'Rugby', 'luta': 'MMA',
    'beisebol': 'Baseball', 'baseball': 'Baseball', 'handebol': 'Handebol',
    'hóquei no gelo': 'Hóquei', 'hoquei no gelo': 'Hóquei', 'hóquei': 'Hóquei',
    'atletismo': 'Atletismo', 'outros esportes': 'Outros',
}


def norm_esporte(esp: str) -> str:
    e = limpa(esp)
    return _ESPORTE_MAP.get(e.lower(), e)


# ---------- tipster: só funde duplicatas de caixa (respeita os nomes) ----------
# "Se tá assim, ta assim" — não renomeia ninguém. O fold é DIRIGIDO PELO DADO:
# variantes que diferem só na caixa (Tp/tp, Peixinho/peixinho) colapsam para a
# forma MAJORITÁRIA daquela planilha. Sem lista fixa → serve qualquer dono.
def construir_fold_tipster(brutos: list[str]) -> dict[str, str]:
    freq = Counter(t for t in brutos if t)
    canon: dict[str, str] = {}
    for lk in {t.lower() for t in freq}:
        variantes = [(freq[t], t) for t in freq if t.lower() == lk]
        canon[lk] = max(variantes)[1]          # mais frequente vence (empate → maior string)
    return canon


# ---------- Status → Resultado (MASTER_RESULTADO §4) ----------
_RESULT_MAP = {
    'ganha': 'W', 'perdida': 'L', 'reembolsada': 'V',
    'meio perdida': 'HL', 'meio ganha': 'HW',
    'pendente': '',            # aposta aberta — sem resultado
}


def norm_resultado(status: str) -> str:
    return _RESULT_MAP.get(limpa(status).lower(), '')


# ---------- mercado → categoria (o "grande de-para": 417 → 27 canônicas) ----------
_MULTI_TIPOS = {'múltipla', 'multipla', 'dupla', 'criar aposta'}


def eh_multipla(tipo: str, esporte: str) -> bool:
    if limpa(tipo).lower() in _MULTI_TIPOS:
        return True
    return limpa(esporte).lower() in ('vários', 'varios')


def categoria(tipo: str, esporte: str, aposta: str, mercado: str) -> str:
    if eh_multipla(tipo, esporte):
        return 'Múltipla'
    esp = norm_esporte(esporte)
    txt = limpa(mercado).lower() + ' || ' + limpa(aposta).lower()

    def has(*ks):
        return any(k in txt for k in ks)

    # Fórmula 1 / Automobilismo: pódio, top-N, qualifying = prop de piloto
    if esp == 'Fórmula 1':
        if has('comparativ', 'head to head', 'h2h'):
            return 'H2H'
        if has('vencedor', 'vencer'):
            return 'ML'
        return 'Player Props'

    if has('escanteio', 'corner'):                              return 'Escanteios'
    if has('cartõe', 'cartoe', 'cartão', 'cartao', 'card'):     return 'Cartões'
    if has('impedimento', 'offside'):                           return 'Impedimentos'
    if has('assistência', 'assistencia', 'assist'):             return 'Assistência'
    if has('desarme', 'tackle'):                                return 'Desarmes'
    if has('no alvo', 'shots on target', 'chutes no gol', 'finalizações no gol'):
        return 'Chutes no Gol'
    if has('chute', 'finalizaç', 'finalizacoes', 'shots'):      return 'Chutes'
    if has('jarda', 'yard'):                                    return 'Jardas'
    if has('ambas', 'ambos marcam', 'ambos os times marcam', 'ambos times marcar',
           'ambos marcarem', 'ambos os time', 'btts', 'both teams'):
        return 'Ambas Marcam'
    if has('dupla chance', 'chance dupla', 'double chance'):     return 'Dupla Chance'
    if has('empate anula', 'draw no bet', 'dnb'):               return 'DNB'
    if esp == 'Dardos' and has('180'):                          return 'Legs'
    if has('best of', 'first to', 'legs', 'leg '):              return 'Legs'
    if has('hat-trick', 'hat trick', 'marcador', 'para marcar', 'de cabeça',
           'a qualquer momento', 'anytime'):
        return 'Anytime'
    if has('total de gols do jogador', 'gols do jogador'):      return 'Player Props'
    # estatísticas de Baseball (bases, strikeouts, hits) — MASTER: Corridas
    if has('total de bases', 'strikeout', 'lançador', 'lancador', 'hits', 'home run',
           'corrida', 'runs', 'baseball'):                      return 'Corridas'
    # marca em ambos os tempos / equipe marca = estatística de equipe
    if has('marca nos dois tempos', 'equipe marca', 'nos dois tempos'):
        return 'Team Props'
    if has('total de jogos', 'games', 'game '):                 return 'Games'
    if has('total de gol', 'gols', 'gol ', 'over/under', 'totais do jogo'): return 'Gols'
    if has('handicap', 'spread', 'linha'):                      return 'Handicap'
    if has('sets', 'set '):                                     return 'Sets'
    if has('duplo-duplo', 'double-double', 'double double'):    return 'Double-Double'
    if has('triplo-duplo', 'triple-double'):                    return 'Triplo-Duplo'
    if has('cestas', 'arremesso', '3 convertidas', 'três pontos', 'defesas', 'goleiro',
           'falta', 'pontos', 'rebote', 'roubos', 'jogador', 'player'):  return 'Player Props'
    if has('moneyline', 'vencedor', 'resultado', '1x2', 'para vencer',
           'match winner', 'partida', 'vencer'):                return 'ML'
    if txt.strip() in ('ml ||', 'ml || ml'):                    return 'ML'
    return 'Outros'


# ---------- data / stake / odd / descrição ----------
_ESP_GENERICO = {'futebol', 'vários', 'varios', 'tênis', 'tenis', 'basquete',
                 'dardos', 'vôlei', 'volei', 'e-sports', 'esports'}


def norm_data(v) -> str:
    """Data do jogo (evento) → DD/MM/AAAA (corta a hora)."""
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def norm_num(v) -> str:
    """Stake/odd → decimal com vírgula (padrão da planilha)."""
    if v is None:
        return ''
    return limpa(v).replace('.', ',')


def norm_desc(evento: str, aposta: str, mercado: str) -> str:
    evt, apo, merc = limpa(evento), limpa(aposta), limpa(mercado)
    base = apo or merc
    if evt and evt.lower() not in _ESP_GENERICO and evt != base and evt not in base:
        return f'{evt} — {base}' if base else evt
    return base or evt


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # colunas: 0 cadastro 1 jogo 2 tipo 3 esporte 4 evento 5 aposta
    #          6 mercado 7 odd 8 stake 9 status 10 casa 11 tipster
    brutas = [r for r in ws.iter_rows(min_row=2, values_only=True)
              if not (r[10] is None and r[7] is None)]     # descarta linha vazia
    fold = construir_fold_tipster([limpa(r[11]) for r in brutas])
    out = []
    for r in brutas:
        tip = limpa(r[11])
        out.append({
            'data': norm_data(r[1]),
            'esporte': norm_esporte(r[3]),
            'tipster': fold.get(tip.lower(), tip),
            'casa': norm_casa(r[10]),
            'parceiro': PARCEIRO,
            'aposta': categoria(r[2], r[3], r[5], r[6]),
            'descricao': norm_desc(r[4], r[5], r[6]),
            'stake': norm_num(r[8]),
            'odd': norm_num(r[7]),
            'resultado': norm_resultado(r[9]),
        })
    return out


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
def _norm_odd(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            _norm_odd(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict], dono: str):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    pares = sorted(set((r['casa'], r['parceiro']) for r in rows))

    registros = []
    for r, sig in zip(rows, sigs):
        estado = 'resolvida' if r['resultado'] in VALID else 'aberta'
        registros.append((
            dono, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None, estado,
            None, None, ORIGEM,                                  # confianca, stake_usd, origem
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem='import'", dono)
                    print(f'  [tentativa {tentativa}] limpou parciais: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    for casa, parceiro in pares:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            dono, casa, parceiro)
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC, id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem='import'
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, dono)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", dono)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", dono)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", dono)
                print(f'\nOK — bilhetes dono={dono}={n} | casas sidebar={nc} | parceiros={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


def _relatorio(rows: list[dict], dono: str):
    print(f'DONO={dono} — linhas a importar: {len(rows)}')
    print('\ncasas:', dict(Counter(r['casa'] for r in rows).most_common()))
    print('\nesporte:', dict(Counter(r['esporte'] for r in rows).most_common()))
    print('\ncategoria (aposta):', dict(Counter(r['aposta'] for r in rows).most_common()))
    print('\nresultado:', dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))
    print('\ntipster:', dict(Counter(r['tipster'] for r in rows).most_common()))
    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} ({dup} desambiguadas por contador)')
    print('\n=== 12 AMOSTRAS (Data | Esporte | Tipster | Casa | Parceiro | Aposta | Descrição | Stake | Odd | Resultado) ===')
    step = max(1, len(rows) // 12)
    for r in rows[::step][:12]:
        print(' | '.join([
            r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
            r['aposta'], r['descricao'][:48], r['stake'], r['odd'], r['resultado'] or '—',
        ]))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dono', required=True)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows, args.dono)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows, args.dono))


if __name__ == '__main__':
    main()
